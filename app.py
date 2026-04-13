import os
import io
import json
import zipfile
import re
import requests
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.utils import get_column_letter
from lxml import etree

app = Flask(__name__)
CORS(app, origins="*")

OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")

# ── COMPANY DATABASE ──────────────────────────────────────────────────────────
COMPANY_DB = """
БАЗА ДАННЫХ КОМПАНИИ ООО «СТРОЙЭКСПЕРТНАДЗОР»:

Полное наименование: Общество с ограниченной ответственностью «Строительная экспертиза и технадзор»
Краткое наименование: ООО «СТРОЙЭКСПЕРТНАДЗОР»
ИНН: 7736557940
КПП: 772901001
ОГРН: 5077746687728
ОКПО: 80751387
ОКАТО: 45268581000
ОКВЭД: 71.2
Дата регистрации: 19.04.2007
Уставный капитал: 10 000 руб.
Юридический адрес: 119415, г. Москва, ул. Удальцова, дом 23, подвал 1, пом. XVА, этаж 1, пом. XXVа
Фактический адрес: 119415, г. Москва, ул. Удальцова, дом 23
Банк: АО «АЛЬФА-БАНК»
БИК: 044525593
Расчётный счёт: 40702810702300004960
Корреспондентский счёт: 30101810200000000593
Телефон: 8 (499) 213-33-13
Email: info@seitn.ru; mail@seitn.ru
Сайт: seitn.ru
Система налогообложения: УСН
Категория МСП: Малое предприятие
Налоговый орган: ИФНС России № 29 по г. Москве
Рег. номер ПФР: 1002601647
Рег. номер ФСС: 1002601647
Генеральный директор: Топчиев Владимир Иванович (ВАЖНО: правильное написание — ТОПчиев, не Толчиев!)
Должность: Генеральный директор
Основание полномочий: Устав
ИНН руководителя: 773601246015
Главный бухгалтер: Гурова Оксана Сергеевна
Среднесписочная численность: 27 чел. (руководители: 4, ИТР/специалисты: 23)
Описание деятельности: Технический заказчик, строительный контроль, обследование и строительная экспертиза

ФИНАНСОВЫЕ ПОКАЗАТЕЛИ (тыс. руб., указывать с пробелами-разделителями):
2025: выручка 197 000 тыс. руб.
2024: выручка 233 000 тыс. руб., чистая прибыль 15 729 тыс. руб.
2023: выручка 127 274 тыс. руб., чистая прибыль 6 992 тыс. руб.
2022: выручка 73 210 тыс. руб.

СРО И ДОПУСКИ:
1. СРО на строительство: Ассоциация «Архитектурное наследие», рег. реестра СРО-С-230-07092010, рег. члена 291116/211, с 29.11.2016, уровень 1 (до 90 млн руб.), включая особо опасные объекты
2. СРО на инженерные изыскания: Ассоциация «Национальный Альянс изыскателей «ГеоЦентр», рег. реестра СРО-И-037-18122012, рег. члена И-037-007736557940-1001, с 25.05.2017, уровень 1 (до 25 млн руб.)
3. СРО на проектирование: Ассоциация проектировщиков «СтройОбъединение», рег. реестра СРО-П-145-04032010, рег. члена П-145-007736557940-0455, с 17.09.2010, уровень 2 (до 50 млн руб.)

ОПЫТ РАБОТ (строительный контроль, последние 3 года):
1. Заказчик: ООО «АШАН» | Объект: Гипермаркеты | Регион: РФ | Период: 2023–н.в. | Сумма: 50 000 000 руб.
2. Заказчик: ФСК Лидер | Объект: Жилой дом | Регион: Москва | Период: 2023–2024 | Сумма: 65 202 300 руб.
3. Заказчик: ООО «Джора» (Azimut) | Объект: Гостиничный комплекс 180 номеров | Регион: Дербент | Период: 2025–н.в. | Сумма: 44 000 000 руб.
4. Заказчик: ООО «Холдинг Строительный Альянс-1» | Объект: Industrial City Block D,E | Регион: Московская область | Период: 2023–2025 | Сумма: 54 000 000 руб.
5. Заказчик: ПАО «МегаФон» | Объект: Центры обработки данных | Регион: СПб, Новосибирск | Период: 2022–н.в. | Сумма: 15 000 000 руб.

Опыт с X5 Group: 121 супермаркет Перекрёсток (2020 г.), Москва, ул. Петрозаводская 24а (обследование, 2021 г.)
"""


# ── XLSX HANDLER ──────────────────────────────────────────────────────────────
def analyze_xlsx(file_bytes):
    """Read xlsx, build cell map and list of empty answer cells with their labels."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # Build complete cell value map
    cell_map = {}
    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value).strip() if cell.value is not None else ""
            cell_map[cell.coordinate] = val

    # Build human-readable structure for AI
    structure_lines = []
    empty_candidates = []  # (coord, label)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        row_parts = []
        has_content = False
        for c in range(1, max_col + 1):
            coord = f"{get_column_letter(c)}{r}"
            val = cell_map.get(coord, "")
            if val:
                has_content = True
            row_parts.append((coord, val))

        if not has_content:
            continue

        # Trim trailing empty
        while row_parts and not row_parts[-1][1]:
            row_parts.pop()

        line = " | ".join(
            f'{coord}="{v[:50]}"' if v else f"{coord}=[ПУСТО]"
            for coord, v in row_parts
        )
        structure_lines.append(line)

    # Find empty answer candidates — first meaningful empty cell per row
    for r in range(1, max_row + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            coord = f"{get_column_letter(c)}{r}"
            row_cells.append((coord, cell_map.get(coord, "")))

        # Find label cells and immediately following empty cells in this row
        for ci, (coord, val) in enumerate(row_cells):
            if val != "":
                continue  # not empty

            # Find nearest label to the left
            label = ""
            for lci in range(ci - 1, -1, -1):
                if row_cells[lci][1]:
                    label = row_cells[lci][1]
                    break

            # If no left label, check up to 3 rows above same column
            if not label:
                for lr in range(r - 1, max(0, r - 4), -1):
                    acoord = f"{get_column_letter(ci + 1)}{lr}"
                    if cell_map.get(acoord, ""):
                        label = cell_map[acoord]
                        break

            # Only add if label found and it's a meaningful answer column
            # Skip cells that are in columns with no adjacent label (e.g. far right decorative cells)
            if label and len(label) > 2:
                # Check this coord isn't already in candidates for this label
                already = any(c == coord for c, l in empty_candidates)
                if not already:
                    empty_candidates.append((coord, label))

    return "\n".join(structure_lines), empty_candidates, cell_map


def fill_xlsx(file_bytes, coord_map):
    """Write values to specified empty cells, return filled xlsx bytes."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # Build set of merged cell top-left corners — these are writable
    # Merged cells: only the top-left cell is writable, others are MergedCell (read-only)
    merged_top_left = set()
    for merged_range in ws.merged_cells.ranges:
        merged_top_left.add(merged_range.coord.split(':')[0])  # top-left coord

    # Build original cell value map for safety check
    orig = {}
    for row in ws.iter_rows():
        for cell in row:
            coord = cell.coordinate
            # MergedCell objects have no .value attribute we can set
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                orig[coord] = "__MERGED__"
            else:
                orig[coord] = str(cell.value).strip() if cell.value is not None else ""

    written = 0
    for coord, value in coord_map.items():
        if not value or value == "[ТРЕБУЕТ УТОЧНЕНИЯ]":
            continue
        cell_orig = orig.get(coord, "")
        # Skip merged non-top-left cells
        if cell_orig == "__MERGED__":
            continue
        # Safety: only write to empty cells
        if cell_orig == "":
            try:
                ws[coord] = value
                written += 1
            except Exception:
                pass  # Skip any problematic cells

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), written


# ── DOCX HANDLER ──────────────────────────────────────────────────────────────
W = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'

def get_cell_text(tc):
    """Extract all text from a table cell."""
    parts = []
    for t in tc.iter(f'{{{W}}}t'):
        if t.text:
            parts.append(t.text)
    return ''.join(parts).strip()


def analyze_docx(file_bytes):
    """Parse docx tables using lxml. Returns structure string, candidates list, tree, all_files."""
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
        all_files = {n: z.read(n) for n in z.namelist()}

    tree = etree.fromstring(all_files['word/document.xml'])
    structure_lines = []
    # candidates: (table_idx, row_idx, label_cell_idx, answer_cell_idx, label_text)
    candidates = []

    for ti, tbl in enumerate(tree.iter(f'{{{W}}}tbl'), 1):
        rows = tbl.findall(f'.//{{{W}}}tr')
        for ri, row in enumerate(rows, 1):
            cells = row.findall(f'{{{W}}}tc')
            cell_texts = [get_cell_text(tc) for tc in cells]

            if not any(cell_texts):
                continue

            parts = ' | '.join(
                f'[{ci}]={repr(t[:40])}' if t else f'[{ci}]=ПУСТО'
                for ci, t in enumerate(cell_texts)
            )
            structure_lines.append(f"T{ti}Строка{ri}: {parts}")

            # Find empty answer cells: label in cell[ci-1], empty in cell[ci]
            for ci, t in enumerate(cell_texts):
                if t == '':
                    label = ''
                    for lci in range(ci - 1, -1, -1):
                        if cell_texts[lci]:
                            label = cell_texts[lci]
                            break
                    if label:
                        candidates.append((ti, ri, ci - 1, ci, label))

    return '\n'.join(structure_lines), candidates, tree, all_files


def fill_docx(file_bytes, fills, tree, all_files):
    """
    fills: list of {table: int, row: int, cell: int, value: str}
    Write values using lxml — correct cell targeting.
    """
    written = 0

    for fill_item in fills:
        ti = fill_item.get('table', 1)
        ri = fill_item['row']
        ci = fill_item['cell']
        value = fill_item['value']
        if not value or value == '[ТРЕБУЕТ УТОЧНЕНИЯ]':
            continue

        # Find the table
        tables = list(tree.iter(f'{{{W}}}tbl'))
        if ti - 1 >= len(tables):
            continue
        tbl = tables[ti - 1]

        rows = tbl.findall(f'.//{{{W}}}tr')
        if ri - 1 >= len(rows):
            continue
        row = rows[ri - 1]

        cells = row.findall(f'{{{W}}}tc')
        if ci >= len(cells):
            continue

        tc = cells[ci]
        # Safety: only write to empty cells
        if get_cell_text(tc):
            continue

        # Find or create paragraph
        para = tc.find(f'{{{W}}}p')
        if para is None:
            para = etree.SubElement(tc, f'{{{W}}}p')

        # Create run with text
        run = etree.SubElement(para, f'{{{W}}}r')
        t_elem = etree.SubElement(run, f'{{{W}}}t')
        t_elem.text = value
        t_elem.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
        written += 1

    # Serialize
    new_xml = etree.tostring(tree, xml_declaration=True, encoding='UTF-8', standalone=True)
    all_files['word/document.xml'] = new_xml

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in all_files.items():
            zout.writestr(name, data)
    out.seek(0)
    return out.read(), written


# ── AI CALL ───────────────────────────────────────────────────────────────────
def call_ai(system_prompt, user_prompt):
    """Call OpenRouter AI and return text response."""
    resp = requests.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://ilovmaks.github.io/tender",
            "X-Title": "Stroyexpertnadzor Tender Assistant",
        },
        json={
            "model": "openrouter/auto",
            "max_tokens": 3000,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        },
        timeout=60,
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


def extract_json(text):
    """Robustly extract JSON object from AI response."""
    # Strip markdown code blocks
    text = re.sub(r"```json\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"```\s*", "", text)
    # Find first { ... }
    m = re.search(r"\{[\s\S]*\}", text)
    if m:
        return json.loads(m.group(0))
    return json.loads(text.strip())


# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "version": "1.0"})


@app.route("/fill", methods=["POST"])
def fill():
    """
    POST /fill
    Form data: file (xlsx or docx)
    Returns: filled file as attachment
    """
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    filename = f.filename or "anketa"
    ext = filename.rsplit(".", 1)[-1].lower()
    file_bytes = f.read()

    if ext not in ("xlsx", "xls", "docx"):
        return jsonify({"error": f"Unsupported format: {ext}"}), 400

    try:
        if ext in ("xlsx", "xls"):
            return fill_xlsx_route(file_bytes, filename)
        else:
            return fill_docx_route(file_bytes, filename)
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


def fill_xlsx_route(file_bytes, filename):
    structure, empty_candidates, cell_map = analyze_xlsx(file_bytes)

    # Build candidate list for AI
    candidates_text = "\n".join(
        f"  {coord} ← метка: \"{label[:70]}\""
        for coord, label in empty_candidates
    )

    system_prompt = f"""Ты заполняешь Excel-анкету данными строительной компании.
Верни ТОЛЬКО валидный JSON объект без markdown и пояснений.
Ключи = адреса ячеек Excel (A1, D4 и т.д.), значения = данные компании.

КРИТИЧЕСКИ ВАЖНО:
1. Заполняй ТОЛЬКО ячейки из списка кандидатов ниже
2. Заполняй ВСЕ поля где есть подходящие данные — не пропускай!
3. Для каждой метки найди значение в базе и запиши
4. Финансы указывай в тыс. руб. с пробелами (197 000, а не 197000)
5. Для таблицы объектов: каждый объект на отдельной строке (B/C/D/E/F)
6. Если поля нет в базе — пропусти ячейку (не включай в JSON)

{COMPANY_DB}"""

    user_prompt = f"""СТРУКТУРА АНКЕТЫ:
{structure}

ЯЧЕЙКИ ДЛЯ ЗАПОЛНЕНИЯ (заполни максимально):
{candidates_text}

Верни JSON со ВСЕМИ заполненными ячейками: {{"D3": "значение", "D4": "значение", ...}}
Не забудь: название компании, адреса, банк, руководитель, СРО, финансы, кадры, объекты опыта."""

    ai_response = call_ai(system_prompt, user_prompt)
    coord_map = extract_json(ai_response)

    # Validate: only allowed coords
    allowed = {coord for coord, _ in empty_candidates}
    coord_map = {k: v for k, v in coord_map.items() if k in allowed}

    filled_bytes, written = fill_xlsx(file_bytes, coord_map)

    base = filename.rsplit(".", 1)[0]
    out_name = f"{base}_заполнено.xlsx"
    return send_file(
        io.BytesIO(filled_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=out_name,
    )


def fill_docx_route(file_bytes, filename):
    structure, candidates, tree, all_files = analyze_docx(file_bytes)

    # candidates: (table_idx, row_idx, label_cell_idx, answer_cell_idx, label_text)
    candidates_text = "\n".join(
        f"  T{ti}Строка{ri} ячейка[{aci}] ← метка: \"{label[:70]}\""
        for ti, ri, lci, aci, label in candidates
    )

    system_prompt = f"""Ты заполняешь Word-анкету (таблица) данными строительной компании.
Верни ТОЛЬКО валидный JSON массив без markdown и пояснений.
Формат: [{{"table": N, "row": N, "cell": N, "value": "..."}}]

Правила:
1. Заполняй ТОЛЬКО ячейки из списка кандидатов
2. Сопоставляй метку с данными компании по смыслу
3. Если данных нет — не включай ячейку
4. Один массив JSON — никакого текста вокруг

{COMPANY_DB}"""

    user_prompt = f"""СТРУКТУРА ТАБЛИЦЫ:
{structure}

ЯЧЕЙКИ ДЛЯ ЗАПОЛНЕНИЯ (ТОЛЬКО эти, точно по адресам):
{candidates_text}

Верни JSON массив: [{{"table": 1, "row": 2, "cell": 2, "value": "ООО Стройэкспертнадзор"}}, ...]"""

    ai_response = call_ai(system_prompt, user_prompt)

    # Parse JSON array
    ai_text = re.sub(r"```json\s*", "", ai_response, flags=re.IGNORECASE)
    ai_text = re.sub(r"```\s*", "", ai_text)
    m = re.search(r"\[[\s\S]*\]", ai_text)
    if m:
        fills = json.loads(m.group(0))
    else:
        fills = json.loads(ai_text.strip())

    # Validate: only allowed cells
    allowed = {(ti, ri, aci) for ti, ri, lci, aci, _ in candidates}
    fills = [
        f for f in fills
        if (f.get("table", 1), f.get("row"), f.get("cell")) in allowed
    ]

    filled_bytes, written = fill_docx(file_bytes, fills, tree, all_files)

    base = filename.rsplit(".", 1)[0]
    out_name = f"{base}_заполнено.docx"
    return send_file(
        io.BytesIO(filled_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        as_attachment=True,
        download_name=out_name,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
